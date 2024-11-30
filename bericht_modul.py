import streamlit as st
import sqlite3
import pandas as pd
from fpdf import FPDF
import openpyxl
from io import BytesIO
from prognose_modul import lade_modell, erstelle_prognosedaten, generiere_prognosen
import os

# Absoluter Pfad zur Datenbank im temporären Verzeichnis
datenbank_pfad = os.path.join('/tmp', 'mathematik_kurs.db')

# Verbindung zur Datenbank
verbindung = sqlite3.connect(datenbank_pfad, check_same_thread=False)

# PDF-Bericht
def generiere_pdf_bericht(teilnehmer, testergebnisse, prognosedaten):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Titel
    pdf.set_font("Arial", size=14, style='B')
    pdf.cell(200, 10, txt=f"Bericht für {teilnehmer['name']}", ln=True, align='C')

    # Teilnehmerdaten
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Teilnehmerdaten:", ln=True)
    pdf.cell(200, 10, txt=f"SV-Nummer: {teilnehmer['sv_nummer']}", ln=True)
    pdf.cell(200, 10, txt=f"Berufswunsch: {teilnehmer['berufswunsch']}", ln=True)
    pdf.cell(200, 10, txt=f"Eintrittsdatum: {teilnehmer['eintrittsdatum']}", ln=True)
    pdf.cell(200, 10, txt=f"Austrittsdatum: {teilnehmer['austrittsdatum']}", ln=True)

    # Testergebnisse
    pdf.cell(200, 10, txt="Testergebnisse:", ln=True)
    for _, row in testergebnisse.iterrows():
        pdf.cell(200, 10, txt=f"Testdatum: {row['test_datum']}, Gesamtprozent: {row['gesamt_prozent']:.2f}%", ln=True)

    # Prognose
    pdf.cell(200, 10, txt="Prognosedaten:", ln=True)
    for index, row in prognosedaten.iterrows():
        pdf.cell(200, 10, txt=f"Tag {row['Tage']}: {row['gesamt_prozent']:.2f}%", ln=True)

    # PDF speichern
    dateiname = f"{teilnehmer['name']}-Bericht.pdf"
    pdf.output(dateiname)
    return dateiname

# Excel-Bericht
def generiere_excel_bericht(teilnehmer, testergebnisse, prognosedaten):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Bericht {teilnehmer['name']}"

    # Teilnehmerdaten
    sheet.append(["Teilnehmerdaten"])
    sheet.append(["Name", teilnehmer['name']])
    sheet.append(["SV-Nummer", teilnehmer['sv_nummer']])
    sheet.append(["Berufswunsch", teilnehmer['berufswunsch']])
    sheet.append(["Eintrittsdatum", teilnehmer['eintrittsdatum']])
    sheet.append(["Austrittsdatum", teilnehmer['austrittsdatum']])

    # Testergebnisse
    sheet.append([])
    sheet.append(["Testergebnisse"])
    sheet.append(["Testdatum", "Gesamtprozent"])
    for _, row in testergebnisse.iterrows():
        sheet.append([row['test_datum'], row['gesamt_prozent']])

    # Prognose
    sheet.append([])
    sheet.append(["Prognosedaten"])
    sheet.append(["Tag", "Gesamtprozent"])
    for _, row in prognosedaten.iterrows():
        sheet.append([row['Tage'], row['gesamt_prozent']])

    # Excel speichern
    dateiname = f"{teilnehmer['name']}-Bericht.xlsx"
    with BytesIO() as b:
        workbook.save(b)
        b.seek(0)
        return b.getvalue()

# Berichtswesen
def berichtswesen():
    st.header("Berichtswesen")

    # Teilnehmerdaten abrufen
    teilnehmer_df = pd.read_sql_query('SELECT * FROM teilnehmer', verbindung)
    if teilnehmer_df.empty:
        st.warning("Keine Teilnehmer vorhanden. Bitte zuerst Teilnehmer hinzufügen.")
        return

    # Teilnehmer auswählen
    teilnehmer_df['Auswahl'] = teilnehmer_df.apply(lambda x: f"{x['name']} (ID: {x['id']})", axis=1)
    ausgewaehlter = st.selectbox("Teilnehmer auswählen", teilnehmer_df['Auswahl'])
    teilnehmer_id = int(ausgewaehlter.split("ID: ")[1].strip(')'))
    teilnehmer = teilnehmer_df[teilnehmer_df['id'] == teilnehmer_id].iloc[0]

    # Testergebnisse abrufen
    testergebnisse_df = pd.read_sql_query(f'SELECT * FROM testergebnisse WHERE teilnehmer_id = {teilnehmer_id}', verbindung)
    if testergebnisse_df.empty:
        st.error("Keine Testergebnisse für diesen Teilnehmer vorhanden.")
        return

    # Prognosedaten abrufen
    modell = lade_modell()
    prognosedaten = None
    if modell:
        daten = erstelle_prognosedaten(teilnehmer_id)
        if daten is not None:
            vorhersagen = generiere_prognosen(modell, daten)
            prognosedaten = vorhersagen[['Tage', 'gesamt_prozent']]

    # PDF-Bericht generieren
    if st.button("PDF-Bericht erstellen"):
        if prognosedaten is not None:
            pdf_dateiname = generiere_pdf_bericht(teilnehmer, testergebnisse_df, prognosedaten)
            with open(pdf_dateiname, "rb") as pdf_file:
                st.download_button(label="PDF herunterladen", data=pdf_file, file_name=pdf_dateiname)
        else:
            st.error("Prognosedaten konnten nicht erstellt werden.")

    # Excel-Bericht generieren
    if st.button("Excel-Bericht erstellen"):
        if prognosedaten is not None:
            excel_inhalt = generiere_excel_bericht(teilnehmer, testergebnisse_df, prognosedaten)
            st.download_button(label="Excel herunterladen", data=excel_inhalt, file_name=f"{teilnehmer['name']}-Bericht.xlsx")
        else:
            st.error("Prognosedaten konnten nicht erstellt werden.")
